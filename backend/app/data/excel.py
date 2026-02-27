"""Excel ingestion utility.

Reads supplier/material/global-context data from a structured xlsx file.
The file must contain three sheets: Suppliers, Materials, Global.
"""

import logging
import os
from pathlib import Path
from typing import Any

logger = logging.getLogger(__name__)

# backend/app/data/excel.py -> parents[2] = backend
DEFAULT_EXCEL_PATH = str(
    Path(__file__).resolve().parents[2] / "data" / "mock_suppliers_demo.xlsx"
)

# ── Column name normalisers ───────────────────────────────────────────

_SUPPLIER_REQUIRED = {"supplier_id", "name"}
_MATERIAL_REQUIRED = {"material_name"}
_GLOBAL_REQUIRED = {"macro_trend"}


def _row_to_dict(headers: list[str], row: tuple) -> dict[str, Any]:
    return {h: (v if v is not None else "") for h, v in zip(headers, row)}


def _read_sheet(wb, sheet_name: str, required_cols: set[str]) -> list[dict]:
    if sheet_name not in wb.sheetnames:
        logger.warning("Sheet '%s' not found in Excel file - skipping.", sheet_name)
        return []
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(h).strip().lower() if h else "" for h in rows[0]]
    missing = required_cols - set(headers)
    if missing:
        raise ValueError(
            f"Sheet '{sheet_name}' is missing required columns: {missing}. "
            f"Found columns: {headers}"
        )
    return [
        _row_to_dict(headers, row)
        for row in rows[1:]
        if any(v is not None for v in row)
    ]


# ── Public loaders ────────────────────────────────────────────────────


def load_suppliers_from_excel(path: str | None = None) -> list[dict]:
    """Return a list of supplier dicts from the Suppliers sheet."""
    import openpyxl

    path = path or DEFAULT_EXCEL_PATH
    if not os.path.exists(path):
        logger.warning("Excel file not found at %s - returning empty suppliers.", path)
        return []
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    rows = _read_sheet(wb, "Suppliers", _SUPPLIER_REQUIRED)
    wb.close()
    logger.info("Loaded %d suppliers from %s", len(rows), path)
    return rows


def load_materials_from_excel(path: str | None = None) -> list[dict]:
    """Return a list of material dicts from the Materials sheet."""
    import openpyxl

    path = path or DEFAULT_EXCEL_PATH
    if not os.path.exists(path):
        logger.warning("Excel file not found at %s - returning empty materials.", path)
        return []
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    rows = _read_sheet(wb, "Materials", _MATERIAL_REQUIRED)
    wb.close()
    logger.info("Loaded %d materials from %s", len(rows), path)
    return rows


def load_global_context_from_excel(path: str | None = None) -> list[dict]:
    """Return a list of global-context trend dicts from the Global sheet."""
    import openpyxl

    path = path or DEFAULT_EXCEL_PATH
    if not os.path.exists(path):
        logger.warning(
            "Excel file not found at %s - returning empty global context.", path
        )
        return []
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    rows = _read_sheet(wb, "Global", _GLOBAL_REQUIRED)
    wb.close()
    logger.info("Loaded %d global context rows from %s", len(rows), path)
    return rows


def load_all_from_excel(path: str | None = None) -> dict[str, list[dict]]:
    """Convenience wrapper: load all three sheets at once."""
    import openpyxl

    path = path or DEFAULT_EXCEL_PATH
    if not os.path.exists(path):
        logger.warning("Excel file not found at %s - returning empty data.", path)
        return {"suppliers": [], "materials": [], "global": []}
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    suppliers = _read_sheet(wb, "Suppliers", _SUPPLIER_REQUIRED)
    materials = _read_sheet(wb, "Materials", _MATERIAL_REQUIRED)
    global_ctx = _read_sheet(wb, "Global", _GLOBAL_REQUIRED)
    wb.close()
    logger.info(
        "Loaded from Excel: %d suppliers, %d materials, %d global rows",
        len(suppliers),
        len(materials),
        len(global_ctx),
    )
    return {"suppliers": suppliers, "materials": materials, "global": global_ctx}
